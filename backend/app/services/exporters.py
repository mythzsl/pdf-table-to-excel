import base64
import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\x00", "").strip()
    return " ".join(text.split())


def normalize_rows(rows: list[list[object]]) -> list[list[str]]:
    width = max((len(row) for row in rows), default=0)
    normalized: list[list[str]] = []

    for row in rows:
        cleaned = [normalize_cell(cell) for cell in row]
        cleaned.extend([""] * (width - len(cleaned)))
        if any(cell for cell in cleaned):
            normalized.append(cleaned)

    return normalized


def make_xlsx_base64(tables: list[dict]) -> str:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    header_fill = PatternFill("solid", fgColor="EAF2FF")

    for position, table in enumerate(tables, start=1):
        title = table.get("name") or f"Table {position}"
        sheet = workbook.create_sheet(title=safe_sheet_title(title, position))
        rows = table["rows"]

        for row in rows:
            sheet.append(row)

        if rows:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill

        for column in sheet.columns:
            max_length = max((len(str(cell.value or "")) for cell in column), default=10)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, 10), 42)

    buffer = BytesIO()
    workbook.save(buffer)
    return base64.b64encode(buffer.getvalue()).decode("ascii")


def make_csv_base64(tables: list[dict]) -> str:
    buffer = StringIO(newline="")
    writer = csv.writer(buffer)

    for index, table in enumerate(tables):
        if index:
            writer.writerow([])
        writer.writerow([table.get("name") or f"Table {index + 1}"])
        writer.writerows(table["rows"])

    return base64.b64encode(buffer.getvalue().encode("utf-8-sig")).decode("ascii")


def safe_sheet_title(title: str, position: int) -> str:
    invalid = set(r"[]:*?/\\")
    cleaned = "".join("_" if character in invalid else character for character in title)
    cleaned = cleaned.strip() or f"Table {position}"
    return cleaned[:31]

